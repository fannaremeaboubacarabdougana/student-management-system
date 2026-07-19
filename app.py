from flask import Flask, render_template, request, redirect, url_for, flash
import csv
from io import StringIO
from flask import Response
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
from reportlab.lib.units import inch
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask_login import (
    login_user,
    logout_user,
    login_required,
    current_user,
)
from sqlalchemy import func
import os
from config import Config
from extensions import db, login_manager
from models import Student, User, ActivityLog, Setting

# Blueprints (we'll enable these later)
# from routes.auth import auth
# from routes.dashboard import dashboard

app = Flask(__name__)
UPLOAD_FOLDER = "static/uploads"

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config.from_object(Config)

# Initialize extensions
db.init_app(app)
login_manager.init_app(app)

login_manager.login_view = "login"


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def log_activity(action):

    if current_user.is_authenticated:
        activity = ActivityLog(username=current_user.username, action=action)

        db.session.add(activity)

        db.session.commit()


@app.route("/")
@login_required
def home():

    # Dashboard Statistics
    total_students = Student.query.count()

    total_programs = db.session.query(
        func.count(func.distinct(Student.program))
    ).scalar()

    average_level = db.session.query(func.avg(Student.level)).scalar()

    highest_level = db.session.query(func.max(Student.level)).scalar()

    lowest_level = db.session.query(func.min(Student.level)).scalar()

    students_with_email = Student.query.filter(Student.email != None).count()

    # Recent Students
    recent_students = Student.query.order_by(Student.id.desc()).limit(5).all()

    # Students by Program (Table)
    program_counts = (
        db.session.query(Student.program, func.count(Student.id))
        .group_by(Student.program)
        .all()
    )

    # -----------------------------
    # Chart Data
    # -----------------------------

    program_data = (
        db.session.query(Student.program, func.count(Student.id))
        .group_by(Student.program)
        .all()
    )

    level_data = (
        db.session.query(Student.level, func.count(Student.id))
        .group_by(Student.level)
        .order_by(Student.level)
        .all()
    )

    # Handle empty database
    if average_level is None:
        average_level = 0

    if highest_level is None:
        highest_level = 0

    if lowest_level is None:
        lowest_level = 0

    return render_template(
        "index.html",
        # Statistics
        total_students=total_students,
        total_programs=total_programs,
        average_level=round(average_level),
        highest_level=highest_level,
        lowest_level=lowest_level,
        students_with_email=students_with_email,
        # Tables
        recent_students=recent_students,
        program_counts=program_counts,
        # Charts
        program_labels=[row[0] for row in program_data],
        program_counts_chart=[row[1] for row in program_data],
        level_labels=[str(row[0]) for row in level_data],
        level_counts=[row[1] for row in level_data],
    )


@app.route("/students")
@login_required
def students():

    search = request.args.get("search", "")
    page = request.args.get("page", 1, type=int)

    query = Student.query

    if search:
        query = query.filter(
            (Student.student_id.contains(search))
            | (Student.first_name.contains(search))
            | (Student.last_name.contains(search))
            | (Student.email.contains(search))
            | (Student.program.contains(search))
        )

    setting = Setting.query.first()

    per_page = setting.students_per_page if setting else 5

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        "students.html",
        students=pagination.items,
        pagination=pagination,
        search=search,
    )


@app.route("/users")
@login_required
def users():

    # Only admins can access this page
    if current_user.role != "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("home"))

    search = request.args.get("search", "")
    page = request.args.get("page", 1, type=int)

    query = User.query

    if search:
        query = query.filter(
            (User.username.contains(search))
            | (User.email.contains(search))
            | (User.role.contains(search))
        )

    pagination = query.paginate(page=page, per_page=5, error_out=False)

    users = User.query.all()

    total_users = len(users)

    admins = User.query.filter_by(role="admin").count()

    staff = User.query.filter_by(role="staff").count()

    return render_template(
        "users.html",
        users=pagination.items,
        pagination=pagination,
        search=search,
        total_users=total_users,
        admins=admins,
        staff=staff,
    )


@app.errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404


@app.errorhandler(500)
def server_error(e):
    return render_template("500.html"), 500


@app.route("/students/<int:id>")
@login_required
def student_profile(id):

    student = Student.query.get_or_404(id)

    return render_template("student_profile.html", student=student)


@app.route("/students/add", methods=["GET", "POST"])
@login_required
def add_student():

    if request.method == "POST":
        student = Student(
            student_id=request.form["student_id"],
            first_name=request.form["first_name"],
            last_name=request.form["last_name"],
            email=request.form["email"],
            phone=request.form["phone"],
            program=request.form["program"],
            level=request.form["level"],
        )

        db.session.add(student)
        db.session.commit()
        log_activity(f"Added student: {student.first_name} {student.last_name}")

        flash("Student added successfully!", "success")

        return redirect(url_for("students"))

    return render_template("add_student.html")


@app.route("/students/edit/<int:id>", methods=["GET", "POST"])
@login_required
def edit_student(id):

    student = Student.query.get_or_404(id)

    if request.method == "POST":
        student.student_id = request.form["student_id"]
        student.first_name = request.form["first_name"]
        student.last_name = request.form["last_name"]
        student.email = request.form["email"]
        student.phone = request.form["phone"]
        student.program = request.form["program"]
        student.level = request.form["level"]

        db.session.commit()
        log_activity(f"Updated student: {student.first_name} {student.last_name}")

        flash("Student updated successfully!", "success")

        return redirect(url_for("students"))

    return render_template(
        "edit_student.html",
        student=student,
    )


@app.route("/students/delete/<int:id>")
@login_required
def delete_student(id):

    if current_user.role != "admin":
        flash(
            "You do not have permission to delete students.",
            "danger",
        )

        return redirect(url_for("students"))

    student = Student.query.get_or_404(id)

    db.session.delete(student)
    student_name = f"{student.first_name} {student.last_name}"
    db.session.commit()
    log_activity(f"Deleted student: {student_name}")

    flash("Student deleted successfully!", "success")

    return redirect(url_for("students"))


@app.route("/students/export/csv")
@login_required
def export_csv():

    output = StringIO()

    writer = csv.writer(output)

    writer.writerow(
        ["Student ID", "First Name", "Last Name", "Email", "Phone", "Program", "Level"]
    )

    students = Student.query.order_by(Student.first_name).all()

    for student in students:
        writer.writerow(
            [
                student.student_id,
                student.first_name,
                student.last_name,
                student.email,
                student.phone,
                student.program,
                student.level,
            ]
        )

    output.seek(0)
    log_activity("Exported students to CSV")

    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=students.csv"},
    )


@app.route("/students/export/pdf")
@login_required
def export_pdf():

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)

    elements = []

    styles = getSampleStyleSheet()

    title = Paragraph(
        "<b>Student Management System</b><br/>Student Report", styles["Title"]
    )

    elements.append(title)

    elements.append(Paragraph("<br/>", styles["Normal"]))

    data = [["Student ID", "Name", "Email", "Program", "Level"]]

    students = Student.query.order_by(Student.first_name).all()

    for student in students:
        data.append(
            [
                student.student_id,
                f"{student.first_name} {student.last_name}",
                student.email,
                student.program,
                student.level,
            ]
        )

    table = Table(data)

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )

    elements.append(table)

    doc.build(elements)

    pdf = buffer.getvalue()

    buffer.close()
    log_activity("Exported students to PDF")

    return Response(
        pdf,
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=students_report.pdf"},
    )


@app.route("/students/export/excel")
@login_required
def export_excel():

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Students"

    headers = [
        "Student ID",
        "First Name",
        "Last Name",
        "Email",
        "Phone",
        "Program",
        "Level",
    ]

    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )

    header_font = Font(color="FFFFFF", bold=True)

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column)

        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

    students = Student.query.order_by(Student.first_name).all()

    row = 2

    for student in students:
        sheet.cell(row=row, column=1).value = student.student_id
        sheet.cell(row=row, column=2).value = student.first_name
        sheet.cell(row=row, column=3).value = student.last_name
        sheet.cell(row=row, column=4).value = student.email
        sheet.cell(row=row, column=5).value = student.phone
        sheet.cell(row=row, column=6).value = student.program
        sheet.cell(row=row, column=7).value = student.level

        row += 1

    for column in sheet.columns:
        max_length = 0

        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass

        sheet.column_dimensions[column_letter].width = max_length + 3

    output = BytesIO()

    workbook.save(output)

    output.seek(0)
    log_activity("Exported students to Excel")

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students.xlsx"},
    )


@app.route("/activity-logs")
@login_required
def activity_logs():

    if current_user.role != "admin":
        flash("Access denied.", "danger")

        return redirect(url_for("home"))

    logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).all()

    return render_template("activity_logs.html", logs=logs)


@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password):
            login_user(user)
            log_activity("Logged in")

            flash("Welcome back!", "success")

            return redirect(url_for("home"))

        flash(
            "Invalid username or password.",
            "danger",
        )

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():

    log_activity("Logged out")

    logout_user()

    flash(
        "Logged out successfully.",
        "success",
    )

    return redirect(url_for("login"))


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():

    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        email = request.form.get("email", "").strip()

        # Update email
        current_user.email = email

        # Change password only if a new password was entered
        if new_password:
            if len(new_password) < 8:
                flash("Password must be at least 8 characters.", "danger")
                return redirect(url_for("profile"))

            if not current_user.check_password(current_password):
                flash("Current password is incorrect.", "danger")
                return redirect(url_for("profile"))

            if new_password != confirm_password:
                flash("Passwords do not match.", "danger")
                return redirect(url_for("profile"))

            current_user.set_password(new_password)

        db.session.commit()

        log_activity("Updated profile")

        flash("Profile updated successfully.", "success")

        return redirect(url_for("profile"))

    return render_template("profile.html")


@app.route("/users/add", methods=["GET", "POST"])
@login_required
def add_user():

    if current_user.role != "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("home"))

    if request.method == "POST":
        username = request.form["username"].strip()
        email = request.form["email"].strip()
        password = request.form["password"]
        role = request.form["role"]

        # Check username
        if User.query.filter_by(username=username).first():
            flash("Username already exists.", "danger")
            return redirect(url_for("add_user"))

        # Check email
        if User.query.filter_by(email=email).first():
            flash("Email already exists.", "danger")
            return redirect(url_for("add_user"))

        user = User(
            username=username,
            email=email,
            role=role,
        )

        user.set_password(password)

        db.session.add(user)
        db.session.commit()

        flash("User created successfully!", "success")

        return redirect(url_for("users"))

    return render_template("add_user.html")


@app.route("/users/edit/<int:id>", methods=["GET", "POST"])
@login_required
def edit_user(id):

    if current_user.role != "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("home"))

    user = User.query.get_or_404(id)

    if request.method == "POST":
        username = request.form["username"].strip()
        email = request.form["email"].strip()
        role = request.form["role"]

        # Username already used by another user
        existing_user = User.query.filter(
            User.username == username, User.id != id
        ).first()

        if existing_user:
            flash("Username already exists.", "danger")
            return redirect(url_for("edit_user", id=id))

        # Email already used by another user
        existing_email = User.query.filter(User.email == email, User.id != id).first()

        if existing_email:
            flash("Email already exists.", "danger")
            return redirect(url_for("edit_user", id=id))

        user.username = username
        user.email = email
        user.role = role

        db.session.commit()

        flash("User updated successfully!", "success")

        return redirect(url_for("users"))

    return render_template("edit_user.html", user=user)


@app.route("/users/delete/<int:id>")
@login_required
def delete_user(id):

    if current_user.role != "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("home"))

    user = User.query.get_or_404(id)

    # Prevent deleting yourself
    if user.id == current_user.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("users"))

    # Log activity
    log = ActivityLog(
        username=current_user.username, action=f"Deleted user '{user.username}'"
    )

    db.session.add(log)

    db.session.delete(user)

    db.session.commit()

    flash("User deleted successfully!", "success")

    return redirect(url_for("users"))


@app.route("/users/reset-password/<int:id>", methods=["GET", "POST"])
@login_required
def reset_user_password(id):

    if current_user.role != "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("users"))

    user = User.query.get_or_404(id)

    if request.method == "POST":
        new_password = request.form["password"]

        user.set_password(new_password)

        db.session.commit()

        flash("Password reset successfully.", "success")

        return redirect(url_for("users"))

    return render_template("reset_user_password.html", user=user)


@app.context_processor
def inject_settings():
    setting = Setting.query.first()

    return dict(setting=setting)


os.makedirs("database", exist_ok=True)

with app.app_context():
    db.create_all()

    # -----------------------------
    # Create default admin account
    # -----------------------------
    admin = User.query.filter_by(username="admin").first()

    if not admin:
        admin = User(
            username="admin",
            email="admin@example.com",
            role="admin",
        )

        admin.set_password("admin123")

        db.session.add(admin)
        db.session.commit()

        print("✓ Default admin account created")

    # -----------------------------
    # Create default system settings
    # -----------------------------
    setting = Setting.query.first()

    if not setting:
        setting = Setting(
            school_name="Student Management System",
            admin_email="admin@example.com",
            students_per_page=5,
            theme="light",
        )

        db.session.add(setting)
        db.session.commit()

        print("✓ Default settings created")


@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():

    if current_user.role != "admin":
        flash("Only administrators can access Settings.", "danger")
        return redirect(url_for("home"))

    setting = Setting.query.first()

    if not setting:
        setting = Setting()
        db.session.add(setting)
        db.session.commit()

    if request.method == "POST":
        setting.school_name = request.form["school_name"]

        setting.admin_email = request.form["admin_email"]

        setting.students_per_page = int(request.form["students_per_page"])

        setting.theme = request.form["theme"]

        file = request.files.get("logo")

        if file and file.filename:
            filename = file.filename

            file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))

            setting.logo = filename

        db.session.commit()

        flash("Settings updated successfully!", "success")

        return redirect(url_for("settings"))

    return render_template("settings.html", setting=setting)


if __name__ == "__main__":
    app.run(debug=True)
